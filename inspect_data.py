import openpyxl
import csv

# Inspect customer_data.xlsx
print("Customer data:")
wb = openpyxl.load_workbook('customer_data.xlsx')
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    print(row)

print("\nLoan data:")
wb2 = openpyxl.load_workbook('loan_data.xlsx')
sheet2 = wb2.active
for row in sheet2.iter_rows(values_only=True):
    print(row)